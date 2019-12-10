#! python3
# spreadsheetToTextFiles.py - Reads the contents of spreadsheet,
# and saves content into text file(s), one row each line and one file each column.

import openpyxl, sys

if len(sys.argv) == 2:

	try:
		wb = openpyxl.load_workbook(sys.argv[1])
		
	except Exception as e:
		print(e)
		
	sheet = wb.active
	
	# Treat each column as a its own new text file.
	for c in range(1, sheet.max_column + 1):

		# First row is treated as a header.
		file_name = sheet.cell(row = 1, column = c).value + '.txt'
		
		text_file = open(file_name, 'w')

		lines = []

		# For all subsequent rows after the header.
		for r in range(2, sheet.max_row + 1):
		
			v = sheet.cell(row = r, column = c).value
		
			# Check if cell has content, then add to list of lines.
			if v is not None:
				lines.append(v)

		text_file.write('\n'.join(lines))
		
		print('Spreadsheet lines from ' + file_name[:-4] + ' successfully saved as ' + file_name)
		
		text_file.close()

else:
	print('You must include a .xlsx file name in your argument.')